from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, FileResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import License, Document
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime


def map_view(request):
    """
    Главная страница с интерактивной картой
    """
    from django.conf import settings
    return render(request, 'licenses/map.html', {
        'yandex_maps_api_key': settings.YANDEX_MAPS_API_KEY
    })


def analytics_view(request):
    """
    Страница аналитики с графиками
    """
    return render(request, 'licenses/analytics.html')


def help_page(request):
    """
    Страница документации и помощи для пользователей
    """
    return render(request, 'licenses/help.html')


def licenses_json(request):
    """
    API endpoint для получения списка лицензий в формате JSON с пагинацией
    """
    licenses = License.objects.all()
    
    # Получаем параметры пагинации
    page_number = request.GET.get('page', 1)
    page_size = int(request.GET.get('page_size', 12))
    
    # Создаем пагинатор
    paginator = Paginator(licenses, page_size)
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    data = []
    for license in page_obj:
        # Проверяем и обновляем статус, если срок истек
        license.update_status_if_expired()

        data.append({
            'id': license.id,
            'license_number': license.license_number,
            'license_type': license.license_type,
            'owner': license.owner,
            'latitude': float(license.latitude) if license.latitude else None,
            'longitude': float(license.longitude) if license.longitude else None,
            'polygon_data': license.polygon_data,
            'region': license.region,
            'area': license.area,
            'issue_date': license.issue_date.strftime('%Y-%m-%d'),
            'expiry_date': license.expiry_date.strftime('%Y-%m-%d') if license.expiry_date else None,
            'mineral_type': license.mineral_type,
            'status': license.status,
            'description': license.description,
        })
    
    # Возвращаем данные с метаинформацией о пагинации
    return JsonResponse({
        'results': data,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
        }
    })


def licenses_all_json(request):
    """
    API endpoint для получения ВСЕХ лицензий без пагинации (для статистики и графиков)
    """
    licenses = License.objects.all()
    data = []
    
    for license in licenses:
        # Проверяем и обновляем статус, если срок истек
        license.update_status_if_expired()

        data.append({
            'id': license.id,
            'license_number': license.license_number,
            'license_type': license.license_type,
            'owner': license.owner,
            'latitude': float(license.latitude) if license.latitude else None,
            'longitude': float(license.longitude) if license.longitude else None,
            'polygon_data': license.polygon_data,
            'region': license.region,
            'area': license.area,
            'issue_date': license.issue_date.strftime('%Y-%m-%d'),
            'expiry_date': license.expiry_date.strftime('%Y-%m-%d') if license.expiry_date else None,
            'mineral_type': license.mineral_type,
            'status': license.status,
            'description': license.description,
        })
    
    return JsonResponse(data, safe=False)


def license_detail(request, license_id):
    """
    Получение детальной информации о лицензии
    """
    license = get_object_or_404(License, id=license_id)
    # Проверяем и обновляем статус, если срок истек
    license.update_status_if_expired()
    documents = license.documents.all()
    
    data = {
        'id': license.id,
        'license_number': license.license_number,
        'license_type': license.license_type,
        'owner': license.owner,
        'latitude': float(license.latitude) if license.latitude else None,
        'longitude': float(license.longitude) if license.longitude else None,
        'polygon_data': license.polygon_data,
        'region': license.region,
        'area': license.area,
        'issue_date': license.issue_date.strftime('%Y-%m-%d'),
        'expiry_date': license.expiry_date.strftime('%Y-%m-%d') if license.expiry_date else None,
        'mineral_type': license.mineral_type,
        'status': license.status,
        'description': license.description,
        'documents': [
            {
                'id': doc.id,
                'title': doc.title,
                'file_type': doc.file_type,
                'file_url': doc.file.url if doc.file else None,
                'uploaded_at': doc.uploaded_at.strftime('%Y-%m-%d %H:%M'),
            }
            for doc in documents
        ]
    }
    
    return JsonResponse(data)


@login_required
def upload_document(request, license_id):
    """
    Загрузка документа для лицензии
    """
    if request.method == 'POST':
        license = get_object_or_404(License, id=license_id)
        
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'Файл не найден'}, status=400)
        
        file = request.FILES['file']
        title = request.POST.get('title', file.name)
        file_type = request.POST.get('file_type', 'other')
        
        document = Document.objects.create(
            license=license,
            title=title,
            file=file,
            file_type=file_type,
            uploaded_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'document_id': document.id,
            'message': 'Документ успешно загружен'
        })
    
    return JsonResponse({'error': 'Метод не поддерживается'}, status=405)


@login_required
def download_document(request, document_id):
    """
    Скачивание документа
    """
    import mimetypes
    import os
    from urllib.parse import quote

    document = get_object_or_404(Document, id=document_id)
    
    if document.file:
        # Получаем оригинальное имя файла
        filename = os.path.basename(document.file.name)
        
        # Определяем MIME-тип по расширению файла
        content_type, _ = mimetypes.guess_type(filename)  
        if not content_type:
            content_type = 'application/octet-stream'
        
        # Открываем файл в бинарном режиме
        file = document.file.open('rb')
        
        response = FileResponse(file, content_type=content_type)
        
        # Правильное формирование Content-Disposition для поддержки кириллицы
        # Используем только один вариант (filename* для Unicode)
        encoded_filename = quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        
        return response
    
    return HttpResponse('Файл не найден', status=404)


def login_view(request):
    """
    Форма входа с поддержкой мультидоменной LDAP аутентификации.
    
    Пользователь может либо выбрать конкретный домен, либо оставить поле пустым
    для автоматического определения домена.
    """
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        domain = request.POST.get('domain', '')
        next_url = request.POST.get('next', '')
        
        # Если домен пустой - передаем None для автоопределения
        domain_param = domain if domain else None
        
        user = authenticate(request, username=username, password=password, domain=domain_param)
        
        if user is not None:
            login(request, user)
            # Перенаправляем на запрошенную страницу или на главную
            if next_url:
                return redirect(next_url)
            return redirect('map')
        else:
            error_message = 'Неверный логин или пароль'
            if domain_param:
                error_message += f' (домен: {domain_param})'
            else:
                error_message += ' (проверены все домены)'
            
            return render(request, 'licenses/login.html', {
                'error': error_message,
                'next': next_url
            })
    
    # GET запрос - проверяем параметр next и добавляем сообщение если нужно
    next_url = request.GET.get('next', '')
    
    # Проверяем, является ли запрошенный URL экспортом или скачиванием
    if next_url:
        if '/api/licenses/export/' in next_url or '/api/documents/' in next_url:
            messages.info(request, 'Для выполнения этого действия необходимо авторизоваться')
    
    return render(request, 'licenses/login.html', {
        'next': next_url
    })


def logout_view(request):
    """
    Выход из системы
    """
    logout(request)
    return redirect('login')


@login_required
def upload_geojson(request):
    """
    Страница для загрузки GeoJSON файлов (только для персонала)
    """
    if not request.user.is_staff:
        return HttpResponse('Доступ запрещён. Только для администраторов.', status=403)
    if request.method == 'POST':
        if 'geojson_file' not in request.FILES:
            return render(request, 'licenses/upload_geojson.html', {
                'error': 'Пожалуйста, выберите GeoJSON файл для загрузки'
            })
        
        geojson_file = request.FILES['geojson_file']
        
        # Проверяем расширение файла
        if not geojson_file.name.endswith('.geojson') and not geojson_file.name.endswith('.json'):
            return render(request, 'licenses/upload_geojson.html', {
                'error': 'Неверный формат файла. Поддерживаются только .geojson и .json файлы'
            })
        
        try:
            # Читаем содержимое файла
            file_content = geojson_file.read().decode('utf-8')
            
            # Импортируем данные
            from .utils import GeoJSONImporter
            importer = GeoJSONImporter()
            result = importer.import_from_file(file_content)
            
            # Формируем сообщение об успехе
            success_message = f"""
                Импорт завершён успешно!
                Создано новых лицензий: {result['imported']}
                Обновлено существующих: {result['updated']}
                Пропущено: {result['skipped']}
                Всего обработано: {result['total']}
            """
            
            return render(request, 'licenses/upload_geojson.html', {
                'success': success_message,
                'result': result
            })
        
        except json.JSONDecodeError:
            return render(request, 'licenses/upload_geojson.html', {
                'error': 'Ошибка: файл не является корректным JSON'
            })
        except Exception as e:
            return render(request, 'licenses/upload_geojson.html', {
                'error': f'Ошибка при обработке файла: {str(e)}'
            })
    
    return render(request, 'licenses/upload_geojson.html')


@login_required
def export_licenses_excel(request):
    """
    Экспорт лицензий в Excel файл с учетом фильтров
    """
    # Получаем параметры фильтрации из GET параметров
    status_filter = request.GET.get('status', '')
    region_filter = request.GET.get('region', '')
    type_filter = request.GET.get('type', '')
    mineral_filter = request.GET.get('mineral', '')
    search_text = request.GET.get('search', '')
    
    # Начинаем с всех лицензий
    licenses = License.objects.all()
    
    # Применяем фильтры
    if status_filter:
        licenses = licenses.filter(status=status_filter)
    if region_filter:
        licenses = licenses.filter(region=region_filter)
    if type_filter:
        licenses = licenses.filter(license_type=type_filter)
    if mineral_filter:
        licenses = licenses.filter(mineral_type=mineral_filter)
    if search_text:
        from django.db.models import Q
        licenses = licenses.filter(
            Q(license_number__icontains=search_text) | 
            Q(owner__icontains=search_text)
        )
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Лицензии"
    
    # Определяем стили
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_style = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Заголовки столбцов
    headers = [
        'Номер лицензии', 'Вид пользования', 'Недропользователь', 
        'Регион', 'Участок недр', 'Дата выдачи', 'Дата окончания',
        'Полезное ископаемое', 'Статус', 'Широта', 'Долгота', 'Описание'
    ]
    
    ws.append(headers)
    
    # Применяем стили к заголовкам
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Заполняем данные
    for license in licenses:
        # Обновляем статус перед экспортом
        license.update_status_if_expired()
        
        status_text = {
            'active': 'Действующая',
            'expired': 'Истекла',
            'suspended': 'Приостановлена',
            'terminated': 'Прекращена'
        }.get(license.status, license.status)
        
        row_data = [
            license.license_number,
            license.license_type,
            license.owner,
            license.region,
            license.area,
            license.issue_date.strftime('%d.%m.%Y') if license.issue_date else '',
            license.expiry_date.strftime('%d.%m.%Y') if license.expiry_date else '',
            license.mineral_type or '',
            status_text,
            float(license.latitude) if license.latitude else '',
            float(license.longitude) if license.longitude else '',
            license.description or ''
        ]
        
        ws.append(row_data)
    
    # Применяем границы ко всем ячейкам с данными
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Автоматически подбираем ширину столбцов
    column_widths = [20, 15, 30, 25, 12, 12, 12, 25, 15, 12, 12, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Фиксируем первую строку
    ws.freeze_panes = 'A2'
    
    # Сохраняем в поток
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Генерируем имя файла с датой и временем
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'licenses_export_{timestamp}.xlsx'
    
    # Возвращаем файл
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def export_licenses_pdf(request):
    """
    Экспорт лицензий в PDF файл с учетом фильтров
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    # Регистрируем DejaVu Sans шрифт для поддержки кириллицы
    # Пробуем разные пути к DejaVu Sans, включая Windows пути
    import os
    import platform
    
    # Пробуем найти шрифты с поддержкой кириллицы в порядке приоритета
    font_candidates = [
        # Arial - отличная поддержка кириллицы, обычно есть в Windows
        ('Arial', [
            'C:/Windows/Fonts/arial.ttf',
            'C:/Windows/Fonts/Arial.ttf',
            '/System/Library/Fonts/Supplemental/Arial.ttf',
            '/Library/Fonts/Arial.ttf',
            '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',  # Linux fallback
        ]),
        # Times New Roman - тоже хорошая поддержка кириллицы
        ('TimesNewRoman', [
            'C:/Windows/Fonts/times.ttf',
            'C:/Windows/Fonts/Times.ttf',
            'C:/Windows/Fonts/timesnr.ttf',
            '/System/Library/Fonts/Supplemental/Times New Roman.ttf',
            '/Library/Fonts/Times New Roman.ttf',
        ]),
        # DejaVu Sans - хороший fallback
        ('DejaVuSans', [
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            '/usr/share/fonts/TTF/DejaVuSans.ttf',
            '/System/Library/Fonts/Supplemental/DejaVuSans.ttf',
            'C:/Windows/Fonts/DejaVuSans.ttf',
        ]),
        # Liberation Sans для Linux
        ('LiberationSans', [
            '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
            '/usr/share/fonts/TTF/LiberationSans-Regular.ttf',
        ]),
    ]
    
    # Добавляем пути из переменных окружения Windows
    if platform.system() == 'Windows':
        windir = os.environ.get('WINDIR', 'C:/Windows')
        # Добавляем Windows пути в начало списков
        for font_name, paths in font_candidates:
            if font_name == 'Arial':
                paths.insert(0, os.path.join(windir, 'Fonts', 'arial.ttf'))
                paths.insert(1, os.path.join(windir, 'Fonts', 'Arial.ttf'))
            elif font_name == 'TimesNewRoman':
                paths.insert(0, os.path.join(windir, 'Fonts', 'times.ttf'))
                paths.insert(1, os.path.join(windir, 'Fonts', 'timesnr.ttf'))
    
    font_registered = False
    font_name = 'Helvetica'  # Fallback по умолчанию
    
    # Пробуем зарегистрировать шрифты в порядке приоритета
    for font_display_name, font_paths in font_candidates:
        for font_path in font_paths:
            try:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont(font_display_name, font_path))
                    font_name = font_display_name
                    font_registered = True
                    break
            except Exception:
                continue
        if font_registered:
            break
    
    # Получаем параметры фильтрации из GET параметров
    status_filter = request.GET.get('status', '')
    region_filter = request.GET.get('region', '')
    type_filter = request.GET.get('type', '')
    mineral_filter = request.GET.get('mineral', '')
    search_text = request.GET.get('search', '')
    
    # Начинаем с всех лицензий
    licenses = License.objects.all()
    
    # Применяем фильтры
    if status_filter:
        licenses = licenses.filter(status=status_filter)
    if region_filter:
        licenses = licenses.filter(region=region_filter)
    if type_filter:
        licenses = licenses.filter(license_type=type_filter)
    if mineral_filter:
        licenses = licenses.filter(mineral_type=mineral_filter)
    if search_text:
        from django.db.models import Q
        licenses = licenses.filter(
            Q(license_number__icontains=search_text) | 
            Q(owner__icontains=search_text)
        )
    
    # Создаем PDF файл
    output = BytesIO()
    
    # Создаем canvas с альбомной ориентацией
    c = canvas.Canvas(output, pagesize=landscape(A4))
    width, height = landscape(A4)
    
    # Заголовок (всегда на русском) - увеличен размер для читаемости
    c.setFont(font_name, 20)
    c.setFillColorRGB(96/255, 165/255, 250/255)  # Синий цвет #60A5FA
    title = 'База лицензий на недропользование'
    c.drawCentredString(width/2, height - 50, title)
    
    # Подзаголовок с датой (всегда на русском) - увеличен размер
    c.setFont(font_name, 12)
    c.setFillColorRGB(107/255, 114/255, 128/255)  # Серый #6B7280
    current_date = datetime.now().strftime('%d.%m.%Y %H:%M')
    subtitle = f'Дата формирования: {current_date}'
    c.drawCentredString(width/2, height - 75, subtitle)
    
    # Начальная позиция для таблицы
    x_start = 30
    
    # Ширины колонок (оптимизированы для альбомной A4)
    col_widths = [25, 80, 65, 150, 90, 65, 65, 100, 90]
    
    # Заголовки таблицы (всегда на русском)
    headers = ['№', 'Номер лицензии', 'Вид', 'Недропользователь', 'Регион', 
               'Дата выдачи', 'Дата окончания', 'Полезное ископаемое', 'Статус']
    
    # Функция для рисования заголовка таблицы
    def draw_table_header(canvas_obj, y_pos):
        canvas_obj.setFillColorRGB(59/255, 130/255, 246/255)  # Синий фон
        canvas_obj.rect(x_start, y_pos - 5, sum(col_widths), 25, fill=1, stroke=0)
        
        canvas_obj.setFillColorRGB(1, 1, 1)  # Белый текст
        canvas_obj.setFont(font_name, 10)  # Увеличен размер для читаемости
        x = x_start + 3
        for i, header in enumerate(headers):
            canvas_obj.drawString(x, y_pos + 5, header)
            x += col_widths[i]
        return y_pos - 30
    
    # Рисуем заголовок на первой странице
    y_position = draw_table_header(c, height - 120)
    row_num = 0
    
    # Данные - экспортируем ВСЕ лицензии - увеличен размер для читаемости
    c.setFont(font_name, 9)
    
    for idx, license in enumerate(licenses, 1):
        if y_position < 50:  # Новая страница если места мало
            c.showPage()
            # Рисуем заголовок на новой странице
            y_position = draw_table_header(c, height - 50)
            c.setFont(font_name, 9)
            row_num = 0
        
        # Обновляем статус перед экспортом
        license.update_status_if_expired()
        
        # Статусы (всегда на русском)
        status_text = {
            'active': 'Действующая',
            'expired': 'Истекла',
            'suspended': 'Приостановлена',
            'terminated': 'Прекращена'
        }.get(license.status, license.status)
        
        # Чередующиеся цвета строк
        if row_num % 2 == 1:
            c.setFillColorRGB(249/255, 250/255, 251/255)  # Светло-серый #F9FAFB
            c.rect(x_start, y_position - 5, sum(col_widths), 20, fill=1, stroke=0)
        
        # Рамка строки
        c.setStrokeColorRGB(229/255, 231/255, 235/255)  # Серая рамка
        c.setLineWidth(0.5)
        c.rect(x_start, y_position - 5, sum(col_widths), 20, fill=0, stroke=1)
        
        c.setFillColorRGB(0, 0, 0)  # Черный текст
        
        row_data = [
            str(idx),
            (license.license_number[:15] + '...') if license.license_number and len(license.license_number) > 15 else (license.license_number or ''),
            (license.license_type[:12] + '...') if license.license_type and len(license.license_type) > 12 else (license.license_type or ''),
            (license.owner[:28] + '...') if license.owner and len(license.owner) > 28 else (license.owner or ''),
            (license.region[:18] + '...') if license.region and len(license.region) > 18 else (license.region or ''),
            license.issue_date.strftime('%d.%m.%Y') if license.issue_date else '',
            license.expiry_date.strftime('%d.%m.%Y') if license.expiry_date else '',
            (license.mineral_type[:18] + '...') if license.mineral_type and len(license.mineral_type) > 18 else (license.mineral_type or ''),
            status_text
        ]
        
        x = x_start + 3
        for i, cell_data in enumerate(row_data):
            c.drawString(x, y_position + 2, cell_data)
            x += col_widths[i]
        
        y_position -= 20
        row_num += 1
    
    # Футер (всегда на русском) - увеличен размер
    c.setFont(font_name, 11)
    c.setFillColorRGB(107/255, 114/255, 128/255)
    footer_text = f'Всего записей: {licenses.count()}'
    c.drawString(x_start, y_position - 20, footer_text)
    
    # Сохраняем PDF
    c.save()
    
    # Получаем данные PDF
    output.seek(0)
    
    # Генерируем имя файла с датой и временем
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'licenses_export_{timestamp}.pdf'
    
    # Возвращаем файл
    response = HttpResponse(output.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response